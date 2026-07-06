"""
Turn a backtest monthly CSV into a readable, formatted Excel workbook.

    python scripts/format_monthly_report.py                     # valuation_momentum
    python scripts/format_monthly_report.py relative_value
    -> exports/backtest_monthly_<basket>.xlsx

Sheet 1 "monthly"  — returns as %, ₹100 curves, drawdowns, turnover; freeze
                     header; green/red scale on the excess-return column.
Sheet 2 "trades"   — per rebalance: names bought / sold (wrapped, wide cols).

Re-run any time after run_backtest.py refreshes the CSV.
"""
from __future__ import annotations
import sys
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

_REPO = Path(__file__).resolve().parents[1]
_BASKET = next((a for a in sys.argv[1:] if not a.startswith("--")), "valuation_momentum")
SRC = _REPO / "exports" / f"backtest_monthly_{_BASKET}.csv"
OUT = _REPO / "exports" / f"backtest_monthly_{_BASKET}.xlsx"

PCT = ["basket_ret", "after_tax_ret", "nifty50_ret", "excess_ret", "tax_drag",
       "basket_drawdown", "nifty50_drawdown", "turnover_two_sided", "est_cost_pct"]
RS = ["cum_basket_rs100", "cum_after_tax_rs100", "cum_nifty50_rs100"]
NICE = {  # column -> (header, width)
    "month": ("Month", 9), "rebalance_date": ("Rebalanced on", 13),
    "basket_ret": ("Basket %", 10), "after_tax_ret": ("After-tax %", 11),
    "nifty50_ret": ("Nifty50 %", 10), "excess_ret": ("Excess %", 10),
    "tax_drag": ("Tax drag %", 10),
    "cum_basket_rs100": ("₹100→Basket", 12), "cum_after_tax_rs100": ("₹100→AfterTax", 13),
    "cum_nifty50_rs100": ("₹100→Nifty50", 12),
    "basket_drawdown": ("Basket DD %", 11), "nifty50_drawdown": ("Nifty DD %", 11),
    "turnover_two_sided": ("Turnover %", 11), "est_cost_pct": ("Cost %", 9),
    "avg_score": ("Avg cheapness (-z)", 16),
    "n_holdings": ("# held", 7), "n_bought": ("# in", 6), "n_sold": ("# out", 6),
}


def main():
    df = pd.read_csv(SRC)
    trades = df[["month", "rebalance_date", "n_bought", "n_sold", "bought", "sold"]].copy()
    monthly = df[[c for c in NICE]].copy()

    def _write(path):
        with pd.ExcelWriter(path, engine="openpyxl") as xw:
            monthly.rename(columns={k: v[0] for k, v in NICE.items()}).to_excel(
                xw, sheet_name="monthly", index=False)
            trades.to_excel(xw, sheet_name="trades", index=False)

            ws = xw.sheets["monthly"]
            ws.freeze_panes = "A2"
            for j, col in enumerate(monthly.columns, start=1):
                letter = get_column_letter(j)
                ws.column_dimensions[letter].width = NICE[col][1] + 2
                ws.cell(row=1, column=j).font = Font(bold=True)
                fmt = "0.00%" if col in PCT else ("#,##0.00" if col in RS or col == "avg_score" else None)
                if fmt:
                    for i in range(2, len(monthly) + 2):
                        ws.cell(row=i, column=j).number_format = fmt
            ex = get_column_letter(monthly.columns.get_loc("excess_ret") + 1)
            ws.conditional_formatting.add(
                f"{ex}2:{ex}{len(monthly) + 1}",
                ColorScaleRule(start_type="num", start_value=-0.03, start_color="F8696B",
                               mid_type="num", mid_value=0, mid_color="FFFFFF",
                               end_type="num", end_value=0.03, end_color="63BE7B"))

            wt = xw.sheets["trades"]
            wt.freeze_panes = "A2"
            for j, w in enumerate([9, 13, 6, 6, 70, 70], start=1):
                wt.column_dimensions[get_column_letter(j)].width = w
                wt.cell(row=1, column=j).font = Font(bold=True)
            for i in range(2, len(trades) + 2):
                for j in (5, 6):
                    wt.cell(row=i, column=j).alignment = Alignment(wrap_text=True, vertical="top")

    try:
        _write(OUT)
        print(f"Wrote {OUT}")
    except PermissionError:
        alt = OUT.with_stem(OUT.stem + "_refreshed")
        _write(alt)
        print(f"⚠ {OUT.name} is OPEN in Excel — could not overwrite it.\n"
              f"  Wrote a fresh copy -> {alt.name}\n"
              f"  Close {OUT.name} in Excel, delete the '~$...' lock file, and re-run to update the original.")


if __name__ == "__main__":
    main()
